import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\DELL.COM\Desktop\Darey\headcount\latest_heascount_from_FA_UPDATED-2.xlsx")
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ===")
    print(f"Dimensions: {ws.dimensions}")
    headers = [cell.value for cell in ws[1]]
    print(f"Headers ({len(headers)}): {headers}")

    if sheet_name == "Headcount":
        print(f"\nMerged cells: {ws.merged_cells}")
        print(f"\nFirst 20 rows cell fill colors:")
        for row_idx in range(1, 21):
            row = ws[row_idx]
            for cell in row[:5]:
                fill = cell.fill
                if fill and fill.fgColor:
                    rgb = fill.fgColor.rgb
                    ftype = fill.fgColor.type
                    if ftype != "none" and rgb not in ("00000000", ""):
                        print(f"  Row {row_idx}, Col {cell.column}: fill={ftype} rgb={rgb} val={cell.value}")
        for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
            print(f"  DataRow: {row[:10]}")
    else:
        print(f"\nAll rows:")
        for row in ws.iter_rows(values_only=True):
            print(f"  {row}")
